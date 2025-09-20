from __future__ import annotations

from io import BytesIO
from typing import Any

from . import register_processor


def _csv_escape(value: Any) -> str:
    if value is None:
        s = ""
    else:
        s = str(value)
    if any(c in s for c in [",", "\n", '"']):
        s = '"' + s.replace('"', '""') + '"'
    return s


def _xlsx_with_pandas(
    data: bytes, *, sheet: str | int | None, max_rows: int
) -> tuple[str, dict[str, Any]]:
    import pandas as pd  # type: ignore

    # Discover sheets
    xls = pd.ExcelFile(BytesIO(data))
    sheet_names: list[str] = list(xls.sheet_names)
    # Choose sheet
    chosen: str
    if isinstance(sheet, str) and sheet in sheet_names:
        chosen = sheet
    elif isinstance(sheet, int) and 0 <= sheet < len(sheet_names):
        chosen = sheet_names[sheet]
    else:
        chosen = sheet_names[0] if sheet_names else "Sheet1"

    df = xls.parse(chosen)
    # Render as CSV text for broad compatibility
    head = df.head(max_rows)
    text = head.to_csv(index=False)
    flags = {
        "kind": "table",
        "rows": int(df.shape[0]),
        "cols": int(df.shape[1]),
        "sheets": sheet_names,
        "sheet_used": chosen,
        "engine": "pandas",
    }
    return text, flags


def _xlsx_with_openpyxl(
    data: bytes, *, sheet: str | int | None, max_rows: int
) -> tuple[str, dict[str, Any]]:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    names = list(wb.sheetnames)

    if isinstance(sheet, str) and sheet in names:
        chosen = sheet
    elif isinstance(sheet, int) and 0 <= sheet < len(names):
        chosen = names[sheet]
    else:
        chosen = names[0] if names else "Sheet1"

    ws = wb[chosen]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > max_rows:
            break
        rows.append([_csv_escape(v) for v in row])
    text = "\n".join([",".join(r) for r in rows])
    flags = {
        "kind": "table",
        "rows": ws.max_row,
        "cols": ws.max_column,
        "sheets": names,
        "sheet_used": chosen,
        "engine": "openpyxl",
    }
    return text, flags


def xlsx_processor(data: bytes, **options: Any) -> dict[str, Any]:
    sheet = options.get("sheet")
    max_rows = int(options.get("max_rows", 200))

    pandas_exc = None
    openpyxl_exc = None

    # Prefer pandas if available
    try:
        import pandas as _  # noqa: F401

        text, flags = _xlsx_with_pandas(data, sheet=sheet, max_rows=max_rows)
        return {"text": text, "images": [], "audio": [], "video": [], "flags": flags}
    except Exception as e:  # pragma: no cover - optional dep
        pandas_exc = str(e)

    # Fallback to openpyxl-only
    try:
        import openpyxl as _  # noqa: F401

        text, flags = _xlsx_with_openpyxl(data, sheet=sheet, max_rows=max_rows)
        return {"text": text, "images": [], "audio": [], "video": [], "flags": flags}
    except Exception as e:  # pragma: no cover - optional dep
        openpyxl_exc = str(e)

    # Neither pandas nor openpyxl available
    return {
        "text": "",
        "images": [],
        "audio": [],
        "video": [],
        "flags": {
            "error": "xlsx processor requires pandas (with openpyxl) or openpyxl",
            "pandas_exc": pandas_exc,
            "openpyxl_exc": openpyxl_exc,
        },
    }


register_processor(".xlsx", xlsx_processor)
